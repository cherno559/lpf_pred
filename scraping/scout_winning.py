"""
sofascore_lpf_generales.py
──────────────────────────
Descarga las estadísticas GENERALES de todos los partidos de una jornada
de la Liga Profesional Argentina y las guarda en un Excel existente.
Ahora soporta Torneo, Temporada y Slug por consola para descargar Playoffs.

Dependencias:
    pip install tls-client openpyxl
"""

import argparse
import time
import sys
import os
import traceback
import tls_client

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────────────────────────
# CONFIGURACIÓN POR DEFECTO
# ──────────────────────────────────────────────────────────────────
TOURNAMENT_ID = 155
SEASON_ID     = 87913
SLUG_RONDA    = ""
BASE_URL      = "https://api.sofascore.com"
SLEEP_REQ     = 2.0

# ──────────────────────────────────────────────────────────────────
# ESTADÍSTICAS ORIGINALES (generales)
# ──────────────────────────────────────────────────────────────────
ESTADISTICAS_CLAVE = {
    "Ball possession":         "Posesión de balón",
    "Expected goals":          "Goles esperados (xG)",
    "Total shots":             "Tiros totales",
    "Shots on target":         "Tiros al arco",
    "Shots off target":        "Tiros afuera",
    "Blocked shots":           "Tiros bloqueados",
    "Corner kicks":            "Córners",
    "Offsides":                "Fueras de juego",
    "Fouls":                   "Faltas",
    "Yellow cards":            "Tarjetas amarillas",
    "Red cards":               "Tarjetas rojas",
    "Passes":                  "Pases totales",
    "Accurate passes":         "Pases precisos",
    "Accurate long balls":     "Balones largos precisos",
    "Accurate crosses":        "Centros precisos",
    "Goalkeeper saves":        "Atajadas del arquero",
    "Tackles":                 "Quites",
    "Interceptions":           "Intercepciones",
    "Clearances":              "Despejes",
    "Big chances":             "Ocasiones claras",
    "Big chances missed":      "Ocasiones claras falladas",
    "Big chances created":     "Ocasiones claras creadas",
    "Passes into final third": "Pases al último tercio",
    "Dribbles":                "Regates intentados",
    "Dribbles succeeded":      "Regates exitosos",
    "Passes accuracy":         "Precisión de pases (%)",
    "Shots inside box":        "Tiros dentro del área",
    "Shots outside box":       "Tiros fuera del área",
    "Expected goals on target":"xG al arco (xGOT)",
    "Goals prevented":         "Goles evitados (arquero)",
    "Duels won":               "Duelos ganados",
    "Aerial duels won":        "Duelos aéreos ganados",
    "Ball recoveries":         "Recuperaciones de balón",
    "Errors leading to shot":  "Errores que generaron tiro",
}

# Métricas de porcentaje que NO deben sumarse entre períodos
# (se recalculan desde los valores absolutos)
METRICAS_PORCENTAJE = {
    "Ball possession",
    "Passes accuracy",
}

# ──────────────────────────────────────────────────────────────────
# MÉTRICAS CALCULADAS
# ──────────────────────────────────────────────────────────────────
def calcular_metricas_derivadas(stats: dict) -> list:
    def n(key, equipo):
        try:
            v = stats.get(key, {}).get(equipo, "0")
            return float(str(v).replace('%', '').strip() or 0)
        except (ValueError, TypeError):
            return 0.0

    derivadas = []

    for eq, label in [("home", "Local"), ("away", "Visitante")]:
        tiros_tot   = n("Total shots", eq)
        tiros_arco  = n("Shots on target", eq)
        conversion  = round(tiros_arco / tiros_tot * 100, 1) if tiros_tot else 0.0
        derivadas.append([f"Precisión de tiros - {label} (%)", conversion, ""])

    for eq, label in [("home", "Local"), ("away", "Visitante")]:
        xg         = n("Expected goals", eq)
        tiros_tot  = n("Total shots", eq)
        xg_por_tiro = round(xg / tiros_tot, 3) if tiros_tot else 0.0
        derivadas.append([f"xG por tiro - {label}", xg_por_tiro, ""])

    for eq, label in [("home", "Local"), ("away", "Visitante")]:
        creadas  = n("Big chances created", eq)
        falladas = n("Big chances missed", eq)
        efic     = round((creadas - falladas) / creadas * 100, 1) if creadas else 0.0
        derivadas.append([f"Eficiencia oportunidades claras - {label} (%)", efic, ""])

    for eq, label in [("home", "Local"), ("away", "Visitante")]:
        intentados = n("Dribbles", eq)
        exitosos   = n("Dribbles succeeded", eq)
        ratio      = round(exitosos / intentados * 100, 1) if intentados else 0.0
        derivadas.append([f"% Regates exitosos - {label}", ratio, ""])

    total_duelos = n("Duels won", "home") + n("Duels won", "away")
    for eq, label in [("home", "Local"), ("away", "Visitante")]:
        ganados = n("Duels won", eq)
        ratio   = round(ganados / total_duelos * 100, 1) if total_duelos else 0.0
        derivadas.append([f"% Duelos ganados - {label}", ratio, ""])

    dif_xg = round(n("Expected goals", "home") - n("Expected goals", "away"), 2)
    derivadas.append(["Diferencial xG (Local - Visitante)", dif_xg, ""])

    for eq, label in [("home", "Local"), ("away", "Visitante")]:
        dentro = n("Shots inside box", eq)
        total  = n("Total shots", eq)
        ratio  = round(dentro / total * 100, 1) if total else 0.0
        derivadas.append([f"% Tiros dentro del área - {label}", ratio, ""])

    return derivadas


EXCEL_DEFAULT = "Fecha_x_fecha_lpf.xlsx"

# ──────────────────────────────────────────────────────────────────
# SESIÓN TLS
# ──────────────────────────────────────────────────────────────────
def build_session() -> tls_client.Session:
    session = tls_client.Session(
        client_identifier="chrome_120",
        random_tls_extension_order=True,
    )
    session.headers.update({
        "accept":                    "application/json, text/plain, */*",
        "accept-encoding":           "gzip, deflate, br",
        "accept-language":           "es-AR,es;q=0.9,en-US;q=0.8,en;q=0.7",
        "cache-control":             "no-cache",
        "dnt":                       "1",
        "origin":                    "https://www.sofascore.com",
        "pragma":                    "no-cache",
        "referer":                   "https://www.sofascore.com/",
        "sec-ch-ua":                 '"Chromium";v="120", "Google Chrome";v="120", "Not-A.Brand";v="99"',
        "sec-ch-ua-mobile":          "?0",
        "sec-ch-ua-platform":        '"Windows"',
        "sec-fetch-dest":            "empty",
        "sec-fetch-mode":            "cors",
        "sec-fetch-site":            "same-site",
        "user-agent":                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                                     "AppleWebKit/537.36 (KHTML, like Gecko) "
                                     "Chrome/120.0.0.0 Safari/537.36",
        "x-requested-with":         "XMLHttpRequest",
    })
    return session

def safe_get(session: tls_client.Session, url: str) -> dict | None:
    for intento in range(3):
        try:
            resp = session.get(url)
            if resp.status_code == 200:
                return resp.json()
            if resp.status_code == 404:
                return None
            if resp.status_code == 403:
                espera = (intento + 1) * 8
                print(f"    ⚠  HTTP 403 (bloqueado) — reintentando en {espera}s…")
                time.sleep(espera)
                continue
            espera = (intento + 1) * 4
            print(f"    ⚠  HTTP {resp.status_code} — reintentando en {espera}s…")
            time.sleep(espera)
        except Exception as e:
            print(f"    ⚠  Excepción: {e}")
            time.sleep(3)
    print(f"    ✗  Falló tras 3 intentos: {url}")
    return None

# ──────────────────────────────────────────────────────────────────
# EXTRACCIÓN DE DATOS
# ──────────────────────────────────────────────────────────────────
def get_events(session, round_number: int) -> list[dict]:
    if SLUG_RONDA:
        url = f"{BASE_URL}/api/v1/unique-tournament/{TOURNAMENT_ID}/season/{SEASON_ID}/events/round/{round_number}/slug/{SLUG_RONDA}"
    else:
        url = f"{BASE_URL}/api/v1/unique-tournament/{TOURNAMENT_ID}/season/{SEASON_ID}/events/round/{round_number}"
        
    data = safe_get(session, url)
    eventos = data.get("events", []) if data else []
    
    # Si la API devuelve los dos torneos juntos (más de 15 partidos), 
    # nos quedamos solo con la segunda mitad (del 16 al 30)
    if len(eventos) > 15:
        return eventos[15:]
        
    return eventos
def extraer_numero(valor) -> float:
    if isinstance(valor, str):
        valor = valor.replace('%', '')
    try:
        return float(valor)
    except (ValueError, TypeError):
        return 0.0


def _acumular_periodos(data: dict) -> dict:
    """
    Construye stats_temporales sumando los períodos 1ST + 2ND (y ET si existe).
    Para métricas de porcentaje (posesión, precisión pases) se promedia en lugar
    de sumar, para no obtener valores absurdos como 101%.
    Devuelve el mismo formato que el camino normal: {nombre: {home: str, away: str}}
    """
    PERIODOS_SUMAR = ('1ST', '2ND', 'ET')
    periodos = [p for p in data['statistics'] if p['period'] in PERIODOS_SUMAR]

    if not periodos:
        # Último recurso: usar el primer período disponible tal cual
        fallback = data['statistics'][0]
        print(f"     ⚠  Usando período fallback único: {fallback['period']}")
        acum = {}
        for grupo in fallback['groups']:
            for item in grupo['statisticsItems']:
                nombre = item.get('name', '')
                if nombre not in acum:
                    acum[nombre] = {'home': item.get('home', '0'), 'away': item.get('away', '0')}
        return acum

    # Acumular sumando
    acum   = {}   # {nombre: {home: float, away: float}}
    conteo = {}   # cuántos períodos aportaron cada métrica (para promediar %)

    for periodo in periodos:
        for grupo in periodo['groups']:
            for item in grupo['statisticsItems']:
                nombre = item.get('name', '')
                if nombre not in acum:
                    acum[nombre]   = {'home': 0.0, 'away': 0.0}
                    conteo[nombre] = 0
                try:
                    acum[nombre]['home'] += float(
                        str(item.get('home', '0')).replace('%', '').strip() or 0
                    )
                    acum[nombre]['away'] += float(
                        str(item.get('away', '0')).replace('%', '').strip() or 0
                    )
                    conteo[nombre] += 1
                except (ValueError, TypeError):
                    pass

    # Promediar las métricas de porcentaje
    resultado = {}
    for nombre, vals in acum.items():
        cnt = conteo[nombre] or 1
        if nombre in METRICAS_PORCENTAJE:
            h = round(vals['home'] / cnt, 1)
            a = round(vals['away'] / cnt, 1)
        else:
            h = vals['home']
            a = vals['away']
        resultado[nombre] = {'home': str(h), 'away': str(a)}

    return resultado


def get_stats_partido(session, event_id: int) -> tuple[list | None, dict]:
    data = safe_get(session, f"{BASE_URL}/api/v1/event/{event_id}/statistics")
    if not data:
        return None, {}

    stats_temporales = {}

    if 'statistics' in data and len(data['statistics']) > 0:

        # ── Intentar período ALL primero (fase regular) ───────────────────
        stats_all = next((p for p in data['statistics'] if p['period'] == 'ALL'), None)

        if stats_all:
            # Camino normal: existe el período consolidado
            for grupo in stats_all['groups']:
                for item in grupo['statisticsItems']:
                    nombre_original = item.get('name', '')
                    if nombre_original not in stats_temporales:
                        stats_temporales[nombre_original] = {
                            'home': item.get('home', '0'),
                            'away': item.get('away', '0')
                        }
        else:
            # ── Fallback: playoffs / tiempo extra / penales ───────────────
            periodos_disp = [p['period'] for p in data['statistics']]
            print(f"     ⚠  Sin período ALL. Períodos disponibles: {periodos_disp}")
            print(f"     →  Acumulando períodos 1ST + 2ND (+ ET si existe)…")
            stats_temporales = _acumular_periodos(data)

    # Construir lista de estadísticas en el orden definido
    estadisticas_partido = []
    for clave_ingles, nombre_espanol in ESTADISTICAS_CLAVE.items():
        if clave_ingles in stats_temporales:
            val_home = stats_temporales[clave_ingles]['home']
            val_away = stats_temporales[clave_ingles]['away']
            estadisticas_partido.append([nombre_espanol, val_home, val_away])

    return (estadisticas_partido if estadisticas_partido else None), stats_temporales

# ──────────────────────────────────────────────────────────────────
# ESTILOS EXCEL
# ──────────────────────────────────────────────────────────────────
def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=11):
    return Font(name="Arial", bold=bold, color=color, size=size)

def _border():
    s = Side(style="thin", color="B0BEC5")
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="center", wrap=False):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

def style(cell, fill, fnt, align=None):
    cell.fill      = fill
    cell.font      = fnt
    cell.border    = _border()
    cell.alignment = align or _align()

F_TITULO   = _fill("0D2B45")
F_HEADER   = _fill("1A3A5C")
F_RESULT   = _fill("2E7D32")
F_EVEN     = _fill("EAF0F6")
F_ODD      = _fill("FFFFFF")
F_DERIVADA = _fill("FFF3CD")
F_DER_HEAD = _fill("E65100")

def write_match_table(ws, start_row: int,
                      home_name: str, away_name: str,
                      goles_local, goles_visitante,
                      estadisticas: list,
                      metricas_derivadas: list) -> int:
    r = start_row
    C = 1

    ws.merge_cells(start_row=r, start_column=C, end_row=r, end_column=C + 2)
    tc = ws.cell(row=r, column=C, value=f"  {home_name}  vs  {away_name}")
    style(tc, F_TITULO, _font(bold=True, color="FFFFFF", size=12))
    r += 1

    for i, label in enumerate(["Métrica", home_name, away_name]):
        c = ws.cell(row=r, column=C + i, value=label)
        style(c, F_HEADER, _font(bold=True, color="FFFFFF"), _align(wrap=True))
    r += 1

    for i, val in enumerate(["Resultado", goles_local, goles_visitante]):
        c = ws.cell(row=r, column=C + i, value=val)
        style(c, F_RESULT, _font(bold=True, color="FFFFFF"))
    r += 1

    for idx, (metrica, val_home, val_away) in enumerate(estadisticas):
        fill_row = F_EVEN if idx % 2 == 0 else F_ODD
        cm = ws.cell(row=r, column=C,     value=metrica)
        ch = ws.cell(row=r, column=C + 1, value=val_home)
        ca = ws.cell(row=r, column=C + 2, value=val_away)
        style(cm, fill_row, _font(), _align(h="left"))
        style(ch, fill_row, _font())
        style(ca, fill_row, _font())
        r += 1

    if metricas_derivadas:
        r += 1

        ws.merge_cells(start_row=r, start_column=C, end_row=r, end_column=C + 2)
        th = ws.cell(row=r, column=C, value="  📊 Métricas derivadas (para modelo predictivo)")
        style(th, F_DER_HEAD, _font(bold=True, color="FFFFFF", size=11))
        r += 1

        for i, label in enumerate(["Métrica calculada", home_name, away_name]):
            c = ws.cell(row=r, column=C + i, value=label)
            style(c, F_HEADER, _font(bold=True, color="FFFFFF"), _align(wrap=True))
        r += 1

        for idx, (metrica, val_home, val_away) in enumerate(metricas_derivadas):
            fill_row = F_DERIVADA if idx % 2 == 0 else F_ODD
            cm = ws.cell(row=r, column=C,     value=metrica)
            ch = ws.cell(row=r, column=C + 1, value=val_home if val_home != "" else "-")
            ca = ws.cell(row=r, column=C + 2, value=val_away if val_away != "" else "-")
            style(cm, fill_row, _font(), _align(h="left"))
            style(ch, fill_row, _font())
            style(ca, fill_row, _font())
            r += 1

    return r

def set_col_widths(ws):
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 24
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        ws.row_dimensions[row[0].row].height = 20

# ──────────────────────────────────────────────────────────────────
# PROCESAMIENTO POR JORNADA
# ──────────────────────────────────────────────────────────────────
def procesar_jornada(session, wb: Workbook, round_number: int):
    sheet_name = f"Fecha {round_number}" if not SLUG_RONDA else f"Fecha {round_number} - {SLUG_RONDA}"

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)

    print(f"\n{'='*54}")
    texto_ronda = f"FECHA {round_number}" if not SLUG_RONDA else f"FECHA {round_number} ({SLUG_RONDA})"
    print(f"  {texto_ronda}")
    print(f"{'='*54}")

    events = get_events(session, round_number)
    if not events:
        print("  ⚠  No se encontraron partidos.")
        ws["A1"] = "Sin partidos disponibles para esta fecha."
        return

    print(f"  {len(events)} partido(s) encontrado(s).\n")
    current_row = 1

    for event in events:
        event_id        = event.get("id")
        home_name       = event.get("homeTeam", {}).get("name", "Local")
        away_name       = event.get("awayTeam", {}).get("name", "Visitante")
        goles_local     = event.get("homeScore", {}).get("current", 0) or 0
        goles_visitante = event.get("awayScore", {}).get("current", 0) or 0

        print(f"  ⚽ {home_name} {goles_local} - {goles_visitante} {away_name}  (id={event_id})")
        time.sleep(SLEEP_REQ)

        print("     Obteniendo estadísticas generales del partido...")
        estadisticas, stats_crudas = get_stats_partido(session, event_id)

        if not estadisticas:
            print("     → Sin estadísticas disponibles, se omite.\n")
            continue

        metricas_derivadas = calcular_metricas_derivadas(stats_crudas)

        next_row = write_match_table(
            ws, current_row,
            home_name, away_name,
            goles_local, goles_visitante,
            estadisticas,
            metricas_derivadas,
        )
        current_row = next_row + 2
        print(f"     ✓ Tabla escrita hasta fila {next_row - 1}\n")

    set_col_widths(ws)

# ──────────────────────────────────────────────────────────────────
# ENTRY POINT
# ──────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description="Scouting Generales – Liga Profesional Argentina (SofaScore)"
    )
    parser.add_argument(
        "--jornada", "-j",
        type=int, nargs="+", required=True,
        help="Número(s) de jornada. Ej: --jornada 5   o   --jornada 1 2 3",
    )
    parser.add_argument(
        "--excel", "-e",
        type=str, default=EXCEL_DEFAULT,
        help=f"Ruta al Excel existente (default: {EXCEL_DEFAULT})",
    )
    parser.add_argument(
        "--torneo", "-t",
        type=int, default=155,
        help="ID del torneo (default: 155 para LPF regular)",
    )
    parser.add_argument(
        "--temporada", "-s",
        type=int, default=87913,
        help="ID de la temporada (default: 87913)",
    )
    parser.add_argument(
        "--slug",
        type=str, default="",
        help="Slug de la ronda para playoffs (ej: round-of-16)",
    )
    args = parser.parse_args()

    global TOURNAMENT_ID, SEASON_ID, SLUG_RONDA
    TOURNAMENT_ID = args.torneo
    SEASON_ID     = args.temporada
    SLUG_RONDA    = args.slug

    ruta_excel = args.excel

    if os.path.exists(ruta_excel):
        print(f"📂 Abriendo Excel existente: {ruta_excel}")
        wb = load_workbook(ruta_excel)
    else:
        print(f"📄 Excel no encontrado — se creará uno nuevo: {ruta_excel}")
        wb = Workbook()
        wb.remove(wb.active)

    print("🌐 Iniciando sesión TLS con SofaScore...")
    session = build_session()

    for i, jornada in enumerate(args.jornada):
        try:
            procesar_jornada(session, wb, jornada)
        except Exception:
            print(f"\n❌ Error procesando Fecha {jornada}:")
            traceback.print_exc()
        if i < len(args.jornada) - 1:
            time.sleep(SLEEP_REQ)

    if not wb.sheetnames:
        print("\n⚠  No se generó ninguna hoja con datos.")
        sys.exit(1)

    wb.save(ruta_excel)
    print(f"\n✅ Excel guardado en: {ruta_excel}")
    print(f"   Hojas disponibles: {', '.join(wb.sheetnames)}")

if __name__ == "__main__":
    main()